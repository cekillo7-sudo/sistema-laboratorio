import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def crear_certificado_excel(datos, nombre_archivo="certificado_generado.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IMPRIMIR"
    ws.views.sheetView[0].showGridLines = True

    # Palette
    NAVY = "1F4E79"
    GOLD = "D9E1F2"
    WHITE = "FFFFFF"
    GRAY_BG = "F2F2F2"
    BORDER_GRAY = "D9D9D9"

    # Fuentes
    font_titulo = Font(name="Calibri", size=14, bold=True, color=NAVY)
    font_header_lab = Font(name="Calibri", size=16, bold=True, color=NAVY)
    font_bold = Font(name="Calibri", size=10, bold=True)
    font_normal = Font(name="Calibri", size=10)
    font_table_head = Font(name="Calibri", size=10, bold=True, color=WHITE)
    font_disclaimer = Font(name="Calibri", size=8, italic=True)

    # Bordes
    thin_border = Border(
        left=Side(style='thin', color=BORDER_GRAY),
        right=Side(style='thin', color=BORDER_GRAY),
        top=Side(style='thin', color=BORDER_GRAY),
        bottom=Side(style='thin', color=BORDER_GRAY)
    )

    # 1. Encabezado de la empresa
    ws["B2"] = "LABORATORIO QUÍMICO METALÚRGICO"
    ws["B2"].font = Font(name="Calibri", size=11, bold=True, color="595959")
    ws["B3"] = "INKA GOLD SILVER S.A.C."
    ws["B3"].font = font_header_lab

    # 2. Título del Certificado
    ws["A5"] = "CERTIFICADO DE ANÁLISIS QUÍMICO N° 000068"
    ws["A5"].font = font_titulo

    # 3. Datos del Cliente / Muestra (Columnas simples: A=Etiqueta, B=Valor)
    info = [
        ("Cliente:", datos.get("cliente", "")),
        ("Tipo de muestra:", datos.get("tipo_muestra", "Mineral")),
        ("Detalle del envase:", datos.get("envase", "1")),
        ("Identificación de la muestra:", datos.get("codigo", "")),
        ("Condición de muestra:", datos.get("condicion", "Sin presinto")),
        ("Fecha de recepción:", datos.get("fecha_recepcion", "")),
        ("Instrumento del análisis:", datos.get("instrumento", "Gravimétrico"))
    ]

    row = 7
    for label, val in info:
        ws.cell(row=row, column=1, value=label).font = font_bold
        ws.cell(row=row, column=2, value=val).font = font_normal
        row += 1

    # 4. Tabla de Resultados (Fila 16 y 17)
    row_table = 16
    headers = [
        ("Descripción de la Muestra", 1),
        ("Porcentaje", 2),
        ("Ley Au (gr/tm)", 3),
        ("Ley Ag (gr/tm)", 4),
        ("Ley Au (OZ/tc)", 5),
        ("Ley Ag (OZ/tc)", 6)
    ]

    # Encabezados de Tabla
    for title, col in headers:
        cell = ws.cell(row=row_table, column=col, value=title)
        cell.font = font_table_head
        cell.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Fila de Valores
    valores = [
        datos.get("codigo", ""),
        f"{datos.get('porcentaje', 100)}%",
        datos.get("ley_au_gr_tm", 0.0),
        "ND",
        datos.get("ley_au_oz_tc", 0.0),
        "ND"
    ]

    for col, val in enumerate(valores, start=1):
        cell = ws.cell(row=row_table + 1, column=col, value=val)
        cell.font = font_bold if col in [3, 5] else font_normal
        cell.alignment = Alignment(horizontal="center" if col > 1 else "left", vertical="center")
        cell.border = thin_border
        if col in [3, 5]:  # Resaltado sutil en Au
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # 5. Fecha de emisión
    ws.cell(row=20, column=1, value="FECHA DE EMISIÓN:").font = font_bold
    ws.cell(row=20, column=2, value=datos.get("fecha_emision", "")).font = font_normal

    # 6. Observaciones y Notas
    notas = [
        "Este informe no debe reproducirse total ni parcialmente sin autorización escrita de Lab. InkaGoldSilver S.A.C.",
        "Los resultados de este certificado solo corresponden a la muestra recibida en nuestra oficina.",
        "Los remanentes de la muestra se guardarán por un periodo máximo de 1 semana."
    ]
    
    r_nota = 22
    for nota in notas:
        c = ws.cell(row=r_nota, column=1, value=nota)
        c.font = font_disclaimer
        r_nota += 1

    # 7. Pie de página / Contacto
    ws.cell(row=27, column=1, value="Av. Los Jazmines S/N & Jr. Los Tulipanes #44 - Llacuabamba").font = font_disclaimer
    ws.cell(row=28, column=1, value="Cel: 913662464 | Email: Lab.inka@gmail.com").font = font_disclaimer

    # Ajustar anchura de columnas automáticamente
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

    wb.save(nombre_archivo)
    print(f"✅ ¡Certificado generado con éxito en '{nombre_archivo}' sin usar plantillas!")

# Prueba del script
if __name__ == "__main__":
    datos_ejemplo = {
        "cliente": "JR.1735.OLENKA",
        "tipo_muestra": "Mineral",
        "envase": "1",
        "codigo": "JR.1735.OLENKA",
        "condicion": "Sin presinto",
        "fecha_recepcion": "31/08/2026",
        "instrumento": "Gravimétrico",
        "porcentaje": 90,
        "ley_au_gr_tm": 3.57,
        "ley_au_oz_tc": 0.10,
        "fecha_emision": "31/08/2026"
    }
    crear_certificado_excel(datos_ejemplo)