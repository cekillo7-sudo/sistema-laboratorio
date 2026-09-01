from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel
import sqlite3
import openpyxl
from datetime import datetime, timedelta
import os
import re
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_NAME = os.path.join(BASE_DIR, "laboratorio_v2.db")

CORREO_EMISOR = "cekillo.7@gmail.com"
PASSWORD_CORREO = "egjr xofj vglb eccw"
CORREO_DESTINO = "cekillo.7@gmail.com" 

class Muestra(BaseModel):
    cliente: str = ""
    codigo: str = ""
    fecha_recepcion: str = ""
    peso_au: float = 0.0
    peso_muestra: float = 10.0
    porcentaje: float = 100.0

def generar_pdf_oficial(datos, pdf_path):
    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=A4,
        rightMargin=40,
        leftMargin=40,
        topMargin=25,
        bottomMargin=25
    )

    elements = []
    styles = getSampleStyleSheet()

    style_title = ParagraphStyle('T1', parent=styles['Normal'], fontSize=13, leading=18, alignment=TA_CENTER, textColor=colors.HexColor('#111827'), fontName='Helvetica-Bold')
    style_label = ParagraphStyle('L1', parent=styles['Normal'], fontSize=10, leading=15, textColor=colors.HexColor('#1F2937'), fontName='Helvetica-Bold')
    style_value = ParagraphStyle('V1', parent=styles['Normal'], fontSize=10, leading=15, textColor=colors.HexColor('#1F2937'), fontName='Helvetica')
    
    style_th = ParagraphStyle('TH', parent=styles['Normal'], fontSize=9, leading=12, alignment=TA_CENTER, textColor=colors.white, fontName='Helvetica-Bold')
    style_th_sub = ParagraphStyle('THS', parent=styles['Normal'], fontSize=9, leading=12, alignment=TA_CENTER, textColor=colors.white, fontName='Helvetica-Bold')
    style_td = ParagraphStyle('TD', parent=styles['Normal'], fontSize=9.5, leading=13, alignment=TA_CENTER, textColor=colors.HexColor('#111827'), fontName='Helvetica')
    style_td_gold = ParagraphStyle('TDG', parent=styles['Normal'], fontSize=9.5, leading=13, alignment=TA_CENTER, textColor=colors.HexColor('#111827'), fontName='Helvetica-Bold')
    
    style_disclaimer = ParagraphStyle('DISC', parent=styles['Normal'], fontSize=8.5, leading=12, textColor=colors.HexColor('#374151'), fontName='Helvetica-Oblique')
    style_footer_text = ParagraphStyle('FT', parent=styles['Normal'], fontSize=8.5, leading=12, textColor=colors.HexColor('#4B5563'), fontName='Helvetica')

    nro_informe = str(datos.get("id_correlativo", 69)).zfill(6)

    # 1. Encabezado Logo
    logo_path = os.path.join(BASE_DIR, "logo.png")
    if os.path.exists(logo_path):
        img_logo = Image(logo_path, width=400, height=100)
        logo_table = Table([[img_logo]], colWidths=[515])
        logo_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ]))
        elements.append(logo_table)
    else:
        elements.append(Paragraph("<b>LABORATORIO QUÍMICO METALÚRGICO</b>", style_title))
        elements.append(Paragraph("<font size=15 color='#C59B27'><b>INKA GOLD SILVER S.A.C.</b></font>", style_title))

    elements.append(Spacer(1, 20))

    # 2. Título Informe (Correlativo)
    elements.append(Paragraph(f"INFORME DE ENSAYO-N° {nro_informe}", style_title))
    elements.append(Spacer(1, 20))

    # 3. Datos del cliente
    info_data = [
        [Paragraph("Cliente:", style_label), Paragraph(str(datos.get("cliente", "")), style_value)],
        [Paragraph("Tipo de muestra:", style_label), Paragraph("Mineral", style_value)],
        [Paragraph("Detalle del envase:", style_label), Paragraph("1", style_value)],
        [Paragraph("Procedencia:", style_label), Paragraph("Llacuabamba", style_value)],
        [Paragraph("Condición de muestra:", style_label), Paragraph("Sin precinto", style_value)],
        [Paragraph("Fecha de recepción:", style_label), Paragraph(str(datos.get("fecha_recepcion", "")), style_value)],
        [Paragraph("Instrumento del análisis:", style_label), Paragraph("Gravimétrico", style_value)]
    ]

    info_table = Table(info_data, colWidths=[140, 375])
    info_table.setStyle(TableStyle([
        ('PADDING', (0,0), (-1,-1), 4),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 25))

    # 4. Tabla de Ensayos
    headers_row1 = [
        Paragraph("Descripcion de la Muestra", style_th),
        Paragraph("Porcentaje", style_th),
        Paragraph("LEYES (gr/tm)", style_th),
        "",
        Paragraph("LEYES (OZ/tc)", style_th),
        ""
    ]
    
    headers_row2 = [
        "",
        "",
        Paragraph("Au (Oro)", style_th_sub),
        Paragraph("Ag (Plata)", style_th_sub),
        Paragraph("Au (Oro)", style_th_sub),
        Paragraph("Ag (Plata)", style_th_sub)
    ]

    row_values = [
        Paragraph(str(datos.get("codigo", "")), style_td),
        Paragraph(f"{int(datos.get('porcentaje', 100))}%", style_td),
        Paragraph(f"{datos.get('ley_au_gr_tm', 0.0):.2f}", style_td_gold),
        Paragraph("ND", style_td),
        Paragraph(f"{datos.get('ley_au_oz_tc', 0.0):.2f}", style_td_gold),
        Paragraph("ND", style_td)
    ]

    res_table = Table([headers_row1, headers_row2, row_values], colWidths=[145, 74, 74, 74, 74, 74])
    res_table.setStyle(TableStyle([
        ('SPAN', (0,0), (0,1)),
        ('SPAN', (1,0), (1,1)),
        ('SPAN', (2,0), (3,0)),
        ('SPAN', (4,0), (5,0)),
        ('BACKGROUND', (0,0), (-1,1), colors.HexColor('#1E293B')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('PADDING', (0,0), (-1,-1), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#94A3B8')),
        ('BACKGROUND', (2,2), (2,2), colors.HexColor('#EAB308')),
        ('BACKGROUND', (4,2), (4,2), colors.HexColor('#EAB308')),
        ('BACKGROUND', (0,2), (1,2), colors.white),
        ('BACKGROUND', (3,2), (3,2), colors.white),
        ('BACKGROUND', (5,2), (5,2), colors.white),
    ]))
    elements.append(res_table)
    elements.append(Spacer(1, 20))

    # 5. Fecha de emisión
    emision_data = [
        [Paragraph("<b>FECHA DE EMISIÓN:</b>", style_label), Paragraph(str(datos.get("fecha_emision", "")), style_value)]
    ]
    emision_table = Table(emision_data, colWidths=[140, 375])
    emision_table.setStyle(TableStyle([
        ('PADDING', (0,0), (-1,-1), 4),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    elements.append(emision_table)
    elements.append(Spacer(1, 20))

    # 6. Disclaimers
    elements.append(Paragraph("Este informe no debe reproducirse total ni parcial sin autorización escrita de Lab.InkaGoldSilver.SAC", style_disclaimer))
    elements.append(Paragraph("Los resultados de este certificado solo corresponden a la muestra recibida en nuestra oficina.", style_disclaimer))
    elements.append(Paragraph("Los remanentes de la muestra se guardaran por un periodo máximo de 1 semana.", style_disclaimer))
    
    # Espacio ampliado para posicionar la firma más abajo de manera correcta
    elements.append(Spacer(1, 55))

    # 7. Sello y firma posicionados con el tamaño de letra reducido (fontSize=6)
    firma_path = os.path.join(BASE_DIR, "firma.png")
    sello_path = os.path.join(BASE_DIR, "sello.png")

    img_firma = Image(firma_path, width=130, height=65) if os.path.exists(firma_path) else Paragraph("", style_value)
    img_sello = Image(sello_path, width=85, height=85) if os.path.exists(sello_path) else Paragraph("", style_value)

    style_subfirma = ParagraphStyle('SF', parent=styles['Normal'], fontSize=6, leading=8, alignment=TA_CENTER, textColor=colors.HexColor('#1F2937'), fontName='Helvetica-Bold')

    firmas_data = [
        [img_firma, img_sello],
        [Paragraph("___________________________________", style_subfirma), ""],
        [Paragraph("<b>INKA GOLD SILVER SAC</b><br/>LABORATORIO QUÍMICO METALÚRGICO", style_subfirma), ""]
    ]

    firmas_data_table = Table(firmas_data, colWidths=[200, 95], hAlign='RIGHT')
    firmas_data_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('SPAN', (1,0), (1,2)),
        ('PADDING', (0,0), (-1,-1), 0),
        ('LEFTPADDING', (1,0), (1,-1), -5),
        ('BOTTOMPADDING', (0,0), (0,0), -15),
        ('TOPPADDING', (0,1), (0,1), 0),
        ('TOPPADDING', (0,2), (0,2), 2),
    ]))
    elements.append(firmas_data_table)
    elements.append(Spacer(1, 20))

    # 8. Pie de página
    elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#CBD5E1'), spaceBefore=5, spaceAfter=8))
    elements.append(Paragraph("Av. Los Jasmines S/N & Jr. Los Tulipanes D44 - Llacuabamba", style_footer_text))
    elements.append(Paragraph("cel : 913662466", style_footer_text))
    elements.append(Paragraph("Email: Lab.inka@gmail.com", style_footer_text))

    doc.build(elements)
    print(f"📄 PDF generado exitosamente: {pdf_path}")

def rellenar_plantilla_excel(datos, archivo_salida):
    plantilla_path = os.path.join(BASE_DIR, "plantilla.xlsx")
    if os.path.exists(plantilla_path):
        try:
            wb = openpyxl.load_workbook(plantilla_path)
            ws = wb.active
            nro_informe = str(datos.get("id_correlativo", 69)).zfill(6)
            ws["B12"] = f"INFORME DE ENSAYO-N° {nro_informe}"
            ws["B15"] = datos.get("cliente", "")
            ws["B20"] = datos.get("fecha_recepcion", "")
            ws["B28"] = datos.get("fecha_emision", "")
            ws["A26"] = datos.get("codigo", "")
            ws["B26"] = f"{int(datos.get('porcentaje', 100))}%"
            ws["C26"] = datos.get("ley_au_gr_tm", 0.0)
            ws["D26"] = "ND"
            ws["E26"] = datos.get("ley_au_oz_tc", 0.0)
            ws["F26"] = "ND"
            wb.save(os.path.join(BASE_DIR, archivo_salida))
        except PermissionError:
            print(f"⚠️ No se pudo guardar {archivo_salida} porque está abierto en Excel. Continuando con PDF...")
        except Exception as e:
            print(f"⚠️ Error en Excel: {e}")

def enviar_reporte_correo(archivo_pdf_path, cliente, nro_informe, ley_au):
    try:
        msg = MIMEMultipart()
        msg['From'] = CORREO_EMISOR
        msg['To'] = CORREO_DESTINO
        msg['Subject'] = f"🧪 Certificado de Ensayo N° {str(nro_informe).zfill(6)} - Cliente: {cliente}"

        cuerpo = f"""
        Hola,

        Se adjunta el certificado oficial de ensayo del laboratorio en formato PDF:

        • Cliente: {cliente}
        • N° de Informe: {str(nro_informe).zfill(6)}
        • Ley Au Calculada: {ley_au} gr/TM

        Saludos,
        Sistema de Control de Laboratorio Químico
        """
        msg.attach(MIMEText(cuerpo, 'plain'))

        with open(archivo_pdf_path, "rb") as f:
            adjunto = MIMEApplication(f.read(), _subtype="pdf")
            adjunto.add_header('Content-Disposition', 'attachment', filename=os.path.basename(archivo_pdf_path))
            msg.attach(adjunto)

        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(CORREO_EMISOR, PASSWORD_CORREO)
        server.send_message(msg)
        server.quit()
        print(f"📧 Correo enviado con el archivo PDF a {CORREO_DESTINO}")
        return True
    except Exception as e:
        print(f"❌ Error al enviar correo: {e}")
        return False

def calcular_fechas(fecha_rec_str):
    dt_rec = None
    formatos = ["%Y-%m-%dT%H:%M", "%d/%m/%Y, %H:%M", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S"]
    
    for fmt in formatos:
        try:
            dt_rec = datetime.strptime(fecha_rec_str.strip(), fmt)
            break
        except Exception:
            pass

    if not dt_rec:
        dt_rec = datetime.now()

    str_recepcion_excel = dt_rec.strftime("%d/%m/%Y")

    if 6 <= dt_rec.hour < 10:
        emision_dt = dt_rec
    elif dt_rec.hour < 6:
        emision_dt = dt_rec
    else:
        emision_dt = dt_rec + timedelta(days=1)

    str_emision_excel = emision_dt.strftime("%d/%m/%Y")
    return str_recepcion_excel, str_emision_excel

@app.on_event("startup")
def init_db():
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS muestras (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo TEXT,
            ley_au REAL,
            cliente TEXT,
            fecha_recepcion TEXT,
            fecha_emision TEXT
        )
    ''')
    conn.commit()
    conn.close()

@app.get("/")
def home():
    index_path = os.path.join(BASE_DIR, "index.html")
    return FileResponse(index_path)

@app.post("/guardar")
async def guardar_muestra(muestra: Muestra):
    try:
        fecha_recepcion, fecha_emision = calcular_fechas(muestra.fecha_recepcion)

        if muestra.peso_muestra > 0:
            ley_au_100 = (muestra.peso_au / muestra.peso_muestra) * 1000000
        else:
            ley_au_100 = 0.0

        factor = muestra.porcentaje / 100.0
        ley_au_final = round(ley_au_100 * factor, 2)
        ley_oz_tc = round(ley_au_final / 34.2857, 2)

        id_informe = 69
        try:
            conn = sqlite3.connect(DB_NAME)
            cursor = conn.cursor()
            cursor.execute("SELECT MAX(id) FROM muestras")
            res = cursor.fetchone()
            max_id = res[0] if res else None
            
            if max_id is None or max_id < 68:
                cursor.execute("INSERT INTO muestras (id, codigo, ley_au, cliente, fecha_recepcion, fecha_emision) VALUES (69, ?, ?, ?, ?, ?)",
                               (muestra.codigo, ley_au_final, muestra.cliente, fecha_recepcion, fecha_emision))
                id_informe = 69
            else:
                cursor.execute("INSERT INTO muestras (codigo, ley_au, cliente, fecha_recepcion, fecha_emision) VALUES (?, ?, ?, ?, ?)", 
                               (muestra.codigo, ley_au_final, muestra.cliente, fecha_recepcion, fecha_emision))
                id_informe = cursor.lastrowid
                
            conn.commit()
            conn.close()
        except Exception as e:
            print("Error en SQLite:", e)

        nro_reporte_archivo = id_informe + 1701
        cliente_limpio = re.sub(r'[^a-zA-Z0-9_-]', '_', muestra.cliente.strip()).upper() if muestra.cliente else "CLIENTE"
        
        nombre_excel = f"REPORTE.{nro_reporte_archivo}.{cliente_limpio}.xlsx"
        nombre_pdf = f"REPORTE.{nro_reporte_archivo}.{cliente_limpio}.pdf"

        datos_excel = {
            "id_correlativo": id_informe,
            "cliente": muestra.cliente,
            "codigo": muestra.codigo,
            "porcentaje": muestra.porcentaje,
            "ley_au_gr_tm": ley_au_final,
            "ley_au_oz_tc": ley_oz_tc,
            "fecha_recepcion": fecha_recepcion,
            "fecha_emision": fecha_emision
        }

        rellenar_plantilla_excel(datos_excel, nombre_excel)

        pdf_full_path = os.path.join(BASE_DIR, nombre_pdf)
        generar_pdf_oficial(datos_excel, pdf_full_path)

        enviar_reporte_correo(pdf_full_path, muestra.cliente, id_informe, ley_au_final)

        return {
            "status": "ok", 
            "mensaje": f"Certificado PDF generado y enviado: {nombre_pdf}", 
            "ley_calculada": ley_au_final, 
            "nro_informe": id_informe,
            "fecha_recepcion": fecha_recepcion,
            "fecha_emision": fecha_emision,
            "archivo": nombre_pdf
        }
    except Exception as e:
        print("❌ Error general en /guardar:", str(e))
        return {
            "status": "error",
            "mensaje": f"Error en el servidor: {str(e)}"
        }

@app.get("/descargar/{nombre_archivo}")
def descargar_reporte(nombre_archivo: str):
    ruta_archivo = os.path.join(BASE_DIR, nombre_archivo)
    if os.path.exists(ruta_archivo):
        media = "application/pdf" if nombre_archivo.endswith(".pdf") else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return FileResponse(ruta_archivo, filename=nombre_archivo, media_type=media)
    return {"status": "error", "mensaje": "Archivo no encontrado"}
@app.get("/health")
def health_check():
    return {"status": "ok"}

if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run("main:app", host="0.0.0.0", port=port)